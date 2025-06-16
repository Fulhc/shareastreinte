from datetime import datetime
import getpass
import os
import io
import openpyxl

from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.folders.folder import Folder
from office365.sharepoint.files.file import File
from office365.runtime.auth.user_credential import UserCredential

# === USER INPUT FOR DATE ===
input_date_str = input("Entrez la date de l'intervention (format YYYY-MM-DD) : ")
try:
    input_date = datetime.strptime(input_date_str, "%Y-%m-%d")
    formatted_date = input_date.strftime("%Y%m%d")
except ValueError:
    print("❌ Format de date invalide. Utilisez le format YYYY-MM-DD.")
    exit(1)

# === DYNAMIC FILENAME ===
FILENAME = f"{formatted_date}_astreinte_Alex.xlsx"

# === LOCAL FILE PATH ===
LOCAL_FILE_PATH = fr"C:\Users\AlexandreBOBEE\OneDrive - FiveNines\Documents\Obsidian\Five9s\🔎⏰Astreintes\{FILENAME}"

# === SHAREPOINT PERSONAL PATH ===
# Replace `john_doe_company_com` with your actual OneDrive ID (based on your email)
SHAREPOINT_FOLDER_URL = "/personal/alexandre_bobee_five9s_fr/Documents/Documents/Obsidian/Five9s/🔎⏰Astreintes"
SHAREPOINT_FILENAME = FILENAME

# === Auto-fill user email based on Windows username and org domain ===
username = getpass.getuser()
email_domain = "five9s.fr"  # ← Replace this with your org domain
user_email = f"{username}@{email_domain}"
print(f"Connexion avec le compte : {user_email}")

user_password = getpass.getpass("Entrez votre mot de passe Microsoft 365 (sera masqué) : ")

CTX = ClientContext("https://fiveninesfr-my.sharepoint.com").with_credentials(
    UserCredential(user_email, user_password)
)

# === Excel File Handling ===
def initialiser_fichier():
    try:
        wb = openpyxl.load_workbook(LOCAL_FILE_PATH)
        if 'Interventions' not in wb.sheetnames:
            ws = wb.create_sheet('Interventions')
            ws.append(['Date', 'Client', 'Tickets Horus', 'Heure de début', 'Heure de fin', 'Durée (HH:MM:SS)'])
            wb.save(LOCAL_FILE_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Interventions'
        ws.append(['Date', 'Client', 'Tickets Horus', 'Heure de début', 'Heure de fin', 'Durée (HH:MM:SS)'])
        wb.save(LOCAL_FILE_PATH)

def debut_intervention():
    heure_debut = datetime.now()
    print(f"Intervention commencée à : {heure_debut.strftime('%Y-%m-%d %H:%M:%S')}")
    return heure_debut

def fin_intervention(heure_debut):
    heure_fin = datetime.now()
    print(f"Intervention terminée à : {heure_fin.strftime('%Y-%m-%d %H:%M:%S')}")
    return heure_fin

def calculer_duree_intervention(heure_debut, heure_fin):
    duree = heure_fin - heure_debut
    heures, reste = divmod(duree.total_seconds(), 3600)
    minutes, secondes = divmod(reste, 60)
    duree_formatee = f"{int(heures):02}:{int(minutes):02}:{int(secondes):02}"
    print(f"Durée de l'intervention : {duree_formatee}")
    return duree_formatee

def enregistrer_intervention(heure_debut, heure_fin, duree_formatee, client, horus_id):
    wb = openpyxl.load_workbook(LOCAL_FILE_PATH)
    ws = wb['Interventions']
    date_intervention = heure_debut.strftime('%Y-%m-%d')
    heure_debut_str = heure_debut.strftime('%H:%M:%S')
    heure_fin_str = heure_fin.strftime('%H:%M:%S')
    ws.append([date_intervention, client, horus_id, heure_debut_str, heure_fin_str, duree_formatee])
    wb.save(LOCAL_FILE_PATH)
    print("✅ Intervention enregistrée dans le fichier Excel.")

# === Upload to SharePoint ===
def upload_to_sharepoint(local_path, ctx, sp_folder_url, sp_filename):
    target_folder = ctx.web.get_folder_by_server_relative_url(sp_folder_url)
    target_folder.upload_file(sp_filename, local_path).execute_query()
    print(f"✅ Fichier téléchargé sur SharePoint : {sp_folder_url}/{sp_filename}")


# === Main Logic ===
def gestion_intervention():
    initialiser_fichier()
    heure_debut = debut_intervention()
    input("Appuyez sur Entrée une fois l'intervention terminée...")
    heure_fin = fin_intervention(heure_debut)
    duree_formatee = calculer_duree_intervention(heure_debut, heure_fin)
    client = input("Quelle est le nom du client : ")
    horus_id = input("Quel est le numéro du ticket Horus : ")
    enregistrer_intervention(heure_debut, heure_fin, duree_formatee, client, horus_id)
    upload_to_sharepoint(LOCAL_FILE_PATH, CTX, SHAREPOINT_FOLDER_URL, SHAREPOINT_FILENAME)

# === Launch ===
gestion_intervention()
